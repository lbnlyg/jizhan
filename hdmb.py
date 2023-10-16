import openpyxl
import os
import tkinter as tk
from tkinter import filedialog

filename =""
def askfile():
    global filename
    # 从本地选择一个文件，并返回文件的目录
    filename = filedialog.askopenfilename(title="请选择基站表格文件",initialdir='.\\')
    if filename != '':
                if filename[-5:] == ".xlsx" or filename[-4:] == ".xls":
                    lb.config(text=  "您选择的基站文件为： "+ filename)
                else:
                    lb.config(text="您选择的文件并非EXCEL表格基站文件")
                    return
    else:
            lb.config(text='您没有选择任何文件')
            return
    
    wb = openpyxl.load_workbook(filename)
    # 显示所有表名
    print("工作表列表:", wb.sheetnames)

    filejztxt = open("jztxt.txt", "w+")
    filejztxt.write('cellID'+"\n")
    # 遍历所有表
    for sheet in wb:
        sheettitle = sheet.title
        print("正在处理：", "="*10, sheet.title, "="*10, "工作表最大列数:",
            wb[sheet.title].max_column, "="*10, "工作表最大行数:", wb[sheet.title].max_row, "="*10)
        #print(sheet.max_column,"="*10,sheet[1])
        cellIDcolumn_letter="-1"
        for cellx in sheet[1]:
            if cellx.value =="CellID":
                print("值",cellx.value,end = "   ")
                print("数字列标",cellx.column,end = "")
                cellIDcolumn_letter = cellx.column_letter
                break
        a="基站类型："+ sheet.title +"  基站数量："  +"基站ID所在列" + cellIDcolumn_letter 
        lb2.config(text=a)
        if cellIDcolumn_letter=="-1":
            print("文件：",filename,"不是基站文件工作表")
            break
        print("  值",cellIDcolumn_letter,end = " \n  ")
        for jizhanID in sheet[cellIDcolumn_letter]:
            if jizhanID.value =="CellID" :
                continue
            filejztxt.write(str(jizhanID.value)+"\n")
            #print("基站ID:",jizhanID.value)
    filejztxt.close


root_window =tk.Tk()
root_window.iconbitmap("tubiao.ico")
# 设置窗口title
root_window.title('政务短信处理助手')
# 设置窗口大小:宽x高,注,此处不能为 "*",必须使用 "x"
root_window.geometry('1000x600')


frame_left_top = tk.LabelFrame(root_window,bg="blue",highlightthickness=10,bd=20)
frame_left_buttom = tk.LabelFrame(root_window,bg="green",highlightthickness=10,bd=20)
frame_right_top = tk.LabelFrame(root_window)
frame_right_buttom = tk.LabelFrame(root_window)

frame_left_top.pack(anchor="nw")
frame_left_buttom.pack(anchor="ne")


# 添加文本内,设置字体的前景色和背景色，和字体类型、大小
text=tk.Label(frame_left_top,text="政务短信基站模板生成",bg="yellow",fg="red",font=('Times', 20, 'bold italic'))
# 将文本内容放置在主窗口内
text.pack()
#添加一个字符输入控件

btn=tk.Button(frame_left_top,text='选择文件',command=askfile)
btn.pack()

lb = tk.Label(frame_left_top,text='',bg='#87CEEB')
lb.pack()

lb2 = tk.Label(frame_left_top,text='',bg='#87CEEB')
lb2.pack()
# 添加按钮，以及按钮的文本，并通过command 参数设置关闭窗口的功能
button=tk.Button(root_window,text="关闭",command=root_window.quit)
# 将按钮放置在主窗口内
button.pack(side="bottom")


#基站模板管理
#c=tk.Label(root_window,text="基站文件为：  "+filename,bg="yellow",fg="red",font=('Times', 20, 'bold italic'))
#c.pack()
#进入主循环，显示主窗口
root_window.mainloop()



        